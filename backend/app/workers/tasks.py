"""
Celery Background Tasks
Task definitions for email, reports, data aggregation, SLA monitoring, etc.
"""

import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from celery import Task
from celery.exceptions import MaxRetriesExceededError

from app.core.performance_config import performance_config
from app.workers.celery_app import celery_app

logger = logging.getLogger(__name__)


class DatabaseTask(Task):
    """
    Base task class with database session management
    """

    _db_session = None

    @property
    def db_session(self):
        """Get or create database session"""
        if self._db_session is None:
            from app.core.database import db_manager

            self._db_session = db_manager.create_session()
        return self._db_session

    def after_return(self, *args, **kwargs):
        """Cleanup after task execution"""
        if self._db_session is not None:
            self._db_session.close()
            self._db_session = None


# Email Tasks
@celery_app.task(
    bind=True,
    base=DatabaseTask,
    max_retries=performance_config.background_jobs.task_max_retries,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_jitter=True,
)
def send_email_task(
    self,
    recipient: str,
    subject: str,
    body: str,
    html_body: Optional[str] = None,
    attachments: Optional[List[Dict]] = None,
):
    """
    Send email asynchronously

    Args:
        recipient: Email recipient
        subject: Email subject
        body: Plain text body
        html_body: HTML body (optional)
        attachments: List of attachments (optional)

    Usage:
        send_email_task.delay(
            recipient="user@example.com",
            subject="Welcome",
            body="Welcome to BARQ!"
        )
    """
    try:
        logger.info(f"Sending email to {recipient}: {subject}")

        import base64
        import os
        import ssl
        from email.mime.application import MIMEApplication
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from smtplib import SMTP, SMTP_SSL

        # Get SMTP configuration from environment
        smtp_host = os.getenv("SMTP_HOST", "")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        smtp_user = os.getenv("SMTP_USER", "")
        smtp_password = os.getenv("SMTP_PASSWORD", "")
        smtp_use_tls = os.getenv("SMTP_USE_TLS", "true").lower() == "true"
        smtp_use_ssl = os.getenv("SMTP_USE_SSL", "false").lower() == "true"
        email_from = os.getenv("EMAIL_FROM", "noreply@barq.com")

        # Create message
        msg = MIMEMultipart("mixed")
        msg["Subject"] = subject
        msg["From"] = email_from
        msg["To"] = recipient

        # Create alternative part for text/html
        msg_alternative = MIMEMultipart("alternative")

        # Add text and HTML parts
        msg_alternative.attach(MIMEText(body, "plain", "utf-8"))
        if html_body:
            msg_alternative.attach(MIMEText(html_body, "html", "utf-8"))

        msg.attach(msg_alternative)

        # Handle attachments if provided
        if attachments:
            for attachment in attachments:
                filename = attachment.get("filename", "attachment")
                content = attachment.get("content")  # Base64 encoded content
                content_type = attachment.get("content_type", "application/octet-stream")

                if content:
                    # Decode base64 content
                    try:
                        file_data = base64.b64decode(content)
                    except Exception:
                        # If not base64, assume it's raw bytes or string
                        file_data = content.encode() if isinstance(content, str) else content

                    # Create attachment
                    part = MIMEApplication(file_data)
                    part.add_header(
                        "Content-Disposition",
                        "attachment",
                        filename=filename
                    )
                    msg.attach(part)
                    logger.debug(f"Attached file: {filename}")

        # Send email if SMTP is configured
        if smtp_host:
            try:
                if smtp_use_ssl:
                    context = ssl.create_default_context()
                    smtp = SMTP_SSL(smtp_host, smtp_port, context=context)
                else:
                    smtp = SMTP(smtp_host, smtp_port)
                    if smtp_use_tls:
                        smtp.starttls()

                if smtp_user and smtp_password:
                    smtp.login(smtp_user, smtp_password)

                smtp.send_message(msg)
                smtp.quit()
                logger.info(f"Email sent successfully to {recipient}")
            except Exception as smtp_error:
                logger.error(f"SMTP error sending email to {recipient}: {smtp_error}")
                raise
        else:
            # Mock mode - log email but don't send
            logger.info(
                f"[MOCK] Email would be sent to {recipient}\n"
                f"Subject: {subject}\n"
                f"Body: {body[:100]}..."
            )

        return {"status": "sent", "recipient": recipient}

    except Exception as e:
        logger.error(f"Failed to send email to {recipient}: {e}")

        # Retry with exponential backoff
        try:
            raise self.retry(exc=e)
        except MaxRetriesExceededError:
            logger.error(f"Max retries exceeded for email to {recipient}")
            return {"status": "failed", "error": str(e)}


@celery_app.task(bind=True, base=DatabaseTask)
def send_bulk_email_task(self, recipients: List[str], subject: str, body: str):
    """
    Send bulk emails

    Args:
        recipients: List of email recipients
        subject: Email subject
        body: Email body

    Usage:
        send_bulk_email_task.delay(
            recipients=["user1@example.com", "user2@example.com"],
            subject="Newsletter",
            body="Latest updates..."
        )
    """
    logger.info(f"Sending bulk email to {len(recipients)} recipients")

    results = []
    for recipient in recipients:
        # Queue individual email tasks
        result = send_email_task.delay(recipient, subject, body)
        results.append(result.id)

    return {"queued_tasks": len(results), "task_ids": results}


# Report Generation Tasks
@celery_app.task(bind=True, base=DatabaseTask)
def generate_report_task(
    self,
    report_type: str,
    organization_id: str,
    start_date: str,
    end_date: str,
    format: str = "pdf",
):
    """
    Generate report asynchronously

    Args:
        report_type: Type of report (performance, financial, etc.)
        organization_id: Organization ID
        start_date: Report start date (ISO format)
        end_date: Report end date (ISO format)
        format: Report format (pdf, xlsx, csv)

    Usage:
        generate_report_task.delay(
            report_type="performance",
            organization_id="org-123",
            start_date="2025-01-01",
            end_date="2025-01-31",
            format="pdf"
        )
    """
    try:
        import csv
        import io
        import json
        import os

        logger.info(
            f"Generating {report_type} report for organization {organization_id} "
            f"({start_date} to {end_date})"
        )

        # Collect report data based on report type
        report_data = _collect_report_data(
            self.db_session,
            report_type=report_type,
            organization_id=organization_id,
            start_date=start_date,
            end_date=end_date,
        )

        # Generate report in requested format
        report_content = None
        content_type = None

        if format == "csv":
            report_content = _generate_csv_report(report_data)
            content_type = "text/csv"
        elif format == "xlsx":
            report_content = _generate_xlsx_report(report_data, report_type)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif format == "pdf":
            report_content = _generate_pdf_report(report_data, report_type, organization_id, start_date, end_date)
            content_type = "application/pdf"
        else:
            # Default to JSON
            report_content = json.dumps(report_data, indent=2, default=str).encode()
            content_type = "application/json"
            format = "json"

        # Save report to storage
        reports_dir = os.getenv("REPORTS_DIR", "/tmp/reports")
        os.makedirs(f"{reports_dir}/{organization_id}", exist_ok=True)
        report_filename = f"{report_type}_{start_date}_{end_date}.{format}"
        report_path = f"{reports_dir}/{organization_id}/{report_filename}"

        with open(report_path, "wb") as f:
            f.write(report_content if isinstance(report_content, bytes) else report_content.encode())

        logger.info(f"Report generated: {report_path}")

        return {
            "status": "completed",
            "report_path": report_path,
            "report_filename": report_filename,
            "content_type": content_type,
            "size_bytes": len(report_content),
            "report_type": report_type,
            "organization_id": organization_id,
            "start_date": start_date,
            "end_date": end_date,
            "format": format,
            "generated_at": datetime.utcnow().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to generate report: {e}")
        raise


@celery_app.task(bind=True, base=DatabaseTask)
def generate_daily_reports_task(self):
    """
    Generate daily reports for all organizations

    Runs daily via Celery Beat
    """
    logger.info("Generating daily reports")

    try:
        from app.models import Organization

        organizations = (
            self.db_session.query(Organization).filter(Organization.is_active == True).all()
        )

        yesterday = (datetime.utcnow() - timedelta(days=1)).date()

        results = []
        for org in organizations:
            # Queue report generation task
            result = generate_report_task.delay(
                report_type="daily_summary",
                organization_id=org.id,
                start_date=yesterday.isoformat(),
                end_date=yesterday.isoformat(),
                format="pdf",
            )
            results.append(result.id)

        logger.info(f"Queued {len(results)} daily report tasks")
        return {"organizations": len(organizations), "task_ids": results}

    except Exception as e:
        logger.error(f"Failed to generate daily reports: {e}")
        raise


# Data Aggregation Tasks
@celery_app.task(bind=True, base=DatabaseTask)
def aggregate_metrics_task(self):
    """
    Aggregate performance metrics

    Runs every 5 minutes via Celery Beat
    """
    logger.info("Aggregating metrics")

    try:
        from app.models.analytics import MetricSnapshot
        from app.utils.analytics import calculate_aggregated_metrics

        # Calculate metrics for all organizations
        metrics = calculate_aggregated_metrics(self.db_session)

        # Store snapshots
        snapshots = []
        for org_id, org_metrics in metrics.items():
            snapshot = MetricSnapshot(
                organization_id=org_id,
                timestamp=datetime.utcnow(),
                metrics=org_metrics,
            )
            snapshots.append(snapshot)

        if snapshots:
            from app.utils.batch import bulk_insert

            bulk_insert(self.db_session, MetricSnapshot, [s.__dict__ for s in snapshots])

        self.db_session.commit()

        logger.info(f"Aggregated metrics for {len(snapshots)} organizations")
        return {"organizations": len(snapshots)}

    except Exception as e:
        logger.error(f"Failed to aggregate metrics: {e}")
        self.db_session.rollback()
        raise


# SLA Monitoring Tasks
@celery_app.task(bind=True, base=DatabaseTask)
def check_sla_compliance_task(self):
    """
    Check SLA compliance for active tickets and deliveries

    Runs every 15 minutes via Celery Beat
    """
    logger.info("Checking SLA compliance")

    try:
        from datetime import datetime

        from app.models.operations import Delivery
        from app.models.support import Ticket

        now = datetime.utcnow()

        # Check ticket SLAs
        overdue_tickets = (
            self.db_session.query(Ticket)
            .filter(
                Ticket.status.in_(["open", "in_progress"]),
                Ticket.sla_due_at < now,
                Ticket.sla_breached == False,
            )
            .all()
        )

        for ticket in overdue_tickets:
            ticket.sla_breached = True
            logger.warning(f"Ticket {ticket.id} breached SLA")

            # TODO: Send notification
            send_email_task.delay(
                recipient=ticket.assigned_to_email,
                subject=f"SLA Breach: Ticket {ticket.number}",
                body=f"Ticket {ticket.number} has breached its SLA.",
            )

        # Check delivery SLAs
        overdue_deliveries = (
            self.db_session.query(Delivery)
            .filter(
                Delivery.status.in_(["pending", "in_transit"]), Delivery.expected_delivery_at < now
            )
            .all()
        )

        for delivery in overdue_deliveries:
            logger.warning(f"Delivery {delivery.id} is overdue")

            # TODO: Send notification

        self.db_session.commit()

        logger.info(
            f"SLA check completed: {len(overdue_tickets)} tickets, "
            f"{len(overdue_deliveries)} deliveries overdue"
        )

        return {
            "overdue_tickets": len(overdue_tickets),
            "overdue_deliveries": len(overdue_deliveries),
        }

    except Exception as e:
        logger.error(f"Failed to check SLA compliance: {e}")
        self.db_session.rollback()
        raise


# Cleanup Tasks
@celery_app.task(bind=True, base=DatabaseTask)
def cleanup_old_data_task(self, days_to_keep: int = 90):
    """
    Cleanup old data (logs, temp files, expired cache, etc.)

    Args:
        days_to_keep: Number of days to keep data

    Runs daily via Celery Beat
    """
    logger.info(f"Cleaning up data older than {days_to_keep} days")

    try:
        from app.models.support import Ticket
        from app.models.workflow import WorkflowHistory

        cutoff_date = datetime.utcnow() - timedelta(days=days_to_keep)

        # Cleanup old workflow history
        deleted_history = (
            self.db_session.query(WorkflowHistory)
            .filter(WorkflowHistory.created_at < cutoff_date)
            .delete()
        )

        # Cleanup resolved tickets older than cutoff
        deleted_tickets = (
            self.db_session.query(Ticket)
            .filter(Ticket.status == "resolved", Ticket.resolved_at < cutoff_date)
            .delete()
        )

        self.db_session.commit()

        logger.info(
            f"Cleanup completed: {deleted_history} history records, "
            f"{deleted_tickets} tickets deleted"
        )

        return {
            "deleted_history": deleted_history,
            "deleted_tickets": deleted_tickets,
        }

    except Exception as e:
        logger.error(f"Failed to cleanup old data: {e}")
        self.db_session.rollback()
        raise


# Cache Warming Tasks
@celery_app.task(bind=True, base=DatabaseTask)
def warm_cache_task(self):
    """
    Warm up cache with frequently accessed data

    Usage:
        warm_cache_task.delay()
    """
    logger.info("Warming cache")

    try:
        from app.core.cache import cache_manager
        from app.models import Organization, User

        # Cache active organizations
        organizations = (
            self.db_session.query(Organization).filter(Organization.is_active == True).all()
        )

        for org in organizations:
            cache_manager.set(
                "organization", org.id, org.to_dict(), ttl=performance_config.cache.organization_ttl
            )

        # Cache active users
        users = self.db_session.query(User).filter(User.is_active == True).limit(1000).all()

        for user in users:
            cache_manager.set(
                "user", user.id, user.to_dict(), ttl=performance_config.cache.user_ttl
            )

        logger.info(f"Cache warmed: {len(organizations)} organizations, {len(users)} users")

        return {
            "organizations": len(organizations),
            "users": len(users),
        }

    except Exception as e:
        logger.error(f"Failed to warm cache: {e}")
        raise


# Report Generation Helper Functions
def _collect_report_data(
    db_session,
    report_type: str,
    organization_id: str,
    start_date: str,
    end_date: str,
) -> Dict[str, Any]:
    """
    Collect data for report based on report type.

    Supported report types:
    - performance: Courier performance metrics
    - financial: Financial summary (payroll, revenue)
    - delivery: Delivery statistics
    - daily_summary: Daily operational summary
    """
    from datetime import datetime as dt

    start = dt.fromisoformat(start_date)
    end = dt.fromisoformat(end_date)

    report_data = {
        "report_type": report_type,
        "organization_id": organization_id,
        "period": {"start_date": start_date, "end_date": end_date},
        "generated_at": datetime.utcnow().isoformat(),
        "data": [],
    }

    try:
        if report_type == "performance":
            from app.models.fleet.courier import Courier
            from app.models.operations.delivery import Delivery, DeliveryStatus

            # Get courier performance data
            couriers = (
                db_session.query(Courier)
                .filter(Courier.organization_id == organization_id)
                .all()
            )

            for courier in couriers:
                deliveries = (
                    db_session.query(Delivery)
                    .filter(
                        Delivery.courier_id == courier.id,
                        Delivery.created_at >= start,
                        Delivery.created_at <= end,
                    )
                    .all()
                )

                completed = sum(1 for d in deliveries if d.status == DeliveryStatus.DELIVERED)
                total = len(deliveries)

                report_data["data"].append({
                    "courier_id": courier.id,
                    "courier_name": courier.full_name,
                    "total_deliveries": total,
                    "completed_deliveries": completed,
                    "success_rate": (completed / total * 100) if total > 0 else 0,
                })

        elif report_type == "financial":
            from app.models.hr.salary import Salary

            # Get salary/financial data
            salaries = (
                db_session.query(Salary)
                .filter(
                    Salary.organization_id == organization_id,
                    Salary.payment_date >= start,
                    Salary.payment_date <= end,
                )
                .all()
            )

            total_payroll = sum(float(s.net_salary or 0) for s in salaries)
            report_data["summary"] = {
                "total_payroll": total_payroll,
                "total_salaries": len(salaries),
            }
            report_data["data"] = [
                {
                    "courier_id": s.courier_id,
                    "month": s.month,
                    "year": s.year,
                    "net_salary": float(s.net_salary or 0),
                    "status": s.status.value if s.status else "unknown",
                }
                for s in salaries
            ]

        elif report_type == "delivery":
            from app.models.operations.delivery import Delivery, DeliveryStatus

            deliveries = (
                db_session.query(Delivery)
                .filter(
                    Delivery.organization_id == organization_id,
                    Delivery.created_at >= start,
                    Delivery.created_at <= end,
                )
                .all()
            )

            status_counts = {}
            for d in deliveries:
                status = d.status.value if d.status else "unknown"
                status_counts[status] = status_counts.get(status, 0) + 1

            report_data["summary"] = {
                "total_deliveries": len(deliveries),
                "by_status": status_counts,
            }
            report_data["data"] = [
                {
                    "delivery_id": d.id,
                    "tracking_number": d.tracking_number,
                    "status": d.status.value if d.status else "unknown",
                    "created_at": d.created_at.isoformat() if d.created_at else None,
                }
                for d in deliveries[:1000]  # Limit to 1000 for performance
            ]

        elif report_type == "daily_summary":
            from app.models.fleet.courier import Courier
            from app.models.operations.delivery import Delivery, DeliveryStatus

            # Daily summary combines multiple metrics
            couriers = (
                db_session.query(Courier)
                .filter(Courier.organization_id == organization_id, Courier.status == "active")
                .count()
            )

            deliveries = (
                db_session.query(Delivery)
                .filter(
                    Delivery.organization_id == organization_id,
                    Delivery.created_at >= start,
                    Delivery.created_at <= end,
                )
                .all()
            )

            completed = sum(1 for d in deliveries if d.status == DeliveryStatus.DELIVERED)
            report_data["summary"] = {
                "active_couriers": couriers,
                "total_deliveries": len(deliveries),
                "completed_deliveries": completed,
                "success_rate": (completed / len(deliveries) * 100) if deliveries else 0,
            }

    except Exception as e:
        logger.error(f"Error collecting report data: {e}")
        report_data["error"] = str(e)

    return report_data


def _generate_csv_report(report_data: Dict[str, Any]) -> bytes:
    """Generate CSV report from data."""
    import csv
    import io

    output = io.StringIO()
    data = report_data.get("data", [])

    if not data:
        output.write("No data available\n")
        return output.getvalue().encode("utf-8")

    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)

    return output.getvalue().encode("utf-8")


def _generate_xlsx_report(report_data: Dict[str, Any], report_type: str) -> bytes:
    """Generate Excel report from data."""
    import io

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace("_", " ").title()

        # Add header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        data = report_data.get("data", [])
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV")
        return _generate_csv_report(report_data)


def _generate_pdf_report(
    report_data: Dict[str, Any],
    report_type: str,
    organization_id: str,
    start_date: str,
    end_date: str,
) -> bytes:
    """Generate PDF report from data."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        import io

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=30,
        )
        title = f"{report_type.replace('_', ' ').title()} Report"
        elements.append(Paragraph(title, title_style))

        # Metadata
        meta = f"Organization: {organization_id}<br/>Period: {start_date} to {end_date}<br/>Generated: {report_data.get('generated_at', 'N/A')}"
        elements.append(Paragraph(meta, styles["Normal"]))
        elements.append(Spacer(1, 20))

        # Summary if available
        summary = report_data.get("summary", {})
        if summary:
            elements.append(Paragraph("Summary", styles["Heading2"]))
            for key, value in summary.items():
                elements.append(Paragraph(f"<b>{key}:</b> {value}", styles["Normal"]))
            elements.append(Spacer(1, 20))

        # Data table
        data = report_data.get("data", [])
        if data:
            elements.append(Paragraph("Details", styles["Heading2"]))
            headers = list(data[0].keys())
            table_data = [headers]
            for row in data[:100]:  # Limit rows for PDF
                table_data.append([str(row.get(h, "")) for h in headers])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

        doc.build(elements)
        return output.getvalue()

    except ImportError:
        logger.warning("reportlab not installed, falling back to plain text")
        # Generate simple text report
        import json
        text = f"Report: {report_type}\n"
        text += f"Organization: {organization_id}\n"
        text += f"Period: {start_date} to {end_date}\n\n"
        text += json.dumps(report_data, indent=2, default=str)
        return text.encode("utf-8")


# Export all tasks
__all__ = [
    "send_email_task",
    "send_bulk_email_task",
    "generate_report_task",
    "generate_daily_reports_task",
    "aggregate_metrics_task",
    "check_sla_compliance_task",
    "cleanup_old_data_task",
    "warm_cache_task",
]
